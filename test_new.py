import numpy as np
import math
from scipy.special import erf, ndtri, erfinv
from openpyxl import load_workbook
class GbE10:
    def __init__(self,worksheet='850S2000'):
        wb = load_workbook('10GbE_inputs.xlsx', data_only=True)
        ws = wb[worksheet]

        # link parameters
        self.br_nominal = ws['C4'].value # nominal baud rate
        self.ber_target = 0 #ws['C5'].value # allowed bit error rate limit
        self.fiber_conn_loss = ws['L7'].value # connector loss
        self.l_target = ws['L3'].value # desired link reach
        self.l_start = ws['L4'].value # minimum link reach
        self.l_inc = ws['L5'].value # link reach increment

        # jitter
        self.dj = ws['G7'].value # deterministic jitter
        self.dcd_dj = ws['G8'].value # duty cycle distortion

        # transmitter performance parameters
        self.tx_oma_min = ws['C8'].value # minimum transmitter optical modulation amplitude
        self.er_dB_min = ws['C9'].value # minimum required extinction ratio
        self.tx_2080_rise = ws['G2'].value # transmitter rise / fall time, 20%-80%
        self.lambda_min = ws['C6'].value # minimum VCSEL wavelength
        self.delta_lambda = ws['C7'].value # VCSEL spectral width
        self.tx_reflection = ws['G12'].value # transmitter reflectance
        self.rin = ws['G4'].value # laser residual intensity noise (RIN)
        self.rin_test_isi = ws['AK7'].value # eye opening, RIN tester
        self.txeye_rx_bw = ws['W5'].value # receiver bandwidth for Tx eye tester
        self.X1 = ws['C12'].value # transmitter eye mask, X1
        self.X2 = ws['C13'].value # transmitter eye mask, X2
        self.Y1 = ws['C14'].value # transmitter eye mask, Y1

        # receiver performance parameters
        self.rx_unstressed_sensitivity = ws['T3'].value # as labeled
        self.rx_bw = ws['T5'].value # receiver bandwidth

        self.rx_reflection = ws['T4'].value # receiver reflectance

        # fiber channel performance parameters
        self.fiber_c_atten = ws['P4'].value # fiber attenuation factor
        self.fiber_s0 = ws['P8'].value # fiber dispersion
        self.fiber_u0 = ws['P7'].value # zero dispersion wavelength for fiber   #Cell Uo -> Check this if not working???
        self.fib_lbwp = ws['P12'].value # fiber length-modal bandwidth product

        # noise penalty parameters
        self.pmn = ws['G13'].value # power penalty for modal noise          #Cell Pmn -> Check this if not working???
        self.ref_nf = ws['L10'].value # noise factor for reflectance
        self.k_mpn = ws['G10'].value # k factor for mode partition noise
        self.sigma_blw = ws['T10'].value # baseline wander noise standard deviation   #Cell SD_blw -> Check this if not working???

        self.eye_time_low = -0.25 #Fixing Value #ws['Z41'].value # minimum eye plot time (UI)
        self.eye_time_high = 1.25 #ws['Z71'].value # maximum eye plot time
        self.eye_time_step = ws['$Y$42'].value # eye plot time increment

        self.time = np.arange(self.eye_time_low, self.eye_time_high,  self.eye_time_step)

        #PRINT

        '''print(self.br_nominal )
        print(self.ber_target) 
        print(self.fiber_conn_loss )
        print( self.l_target)
        print(self.l_start )
        print(self.l_inc)

        # jitter
        print(self.dj)
        print(self.dcd_dj )

        # transmitter performance parameters
        print(self.tx_oma_min)
        print(self.er_dB_min)
        print(self.tx_2080_rise)
        print(self.lambda_min)
        print(self.delta_lambda)
        print(self.tx_reflection )
        print(self.rin )
        print(self.rin_test_isi)
        print(self.txeye_rx_bw)
        print(self.X1 )
        print(self.X2 )
        print(self.Y1 )

        # receiver performance parameters
        print(self.rx_unstressed_sensitivity )
        print(self.rx_bw )
        print(self.rx_reflection)
        
        # fiber channel performance parameters
        print(self.fiber_c_atten )
        print(self.fiber_s0)
        print(self.fiber_u0 )
        print(self.fib_lbwp )

        # noise penalty parameters
        print(self.pmn )
        print(self.ref_nf)
        print(self.k_mpn)
        print(self.sigma_blw)
        print(self.eye_time_low )
        print(self.eye_time_high )
        print(self.eye_time_step )''' 

        self.Q = 7.04



        self.preliminary_calc()
        self.fiber_channel_calc()
        self.risetime_calc()
        self.isi_calc()
        self.penalty_calc()
        self.eye_calc()

        # end of GbE10.__init__
    #======================================================================+


    def preliminary_calc(self):

        # calculate Q from target BER; see equation B.41 in FC-MSQS-2
        #self.Q = -ndtri(self.ber_target)
        self.Q = 7.04
        #print('ber_target : ', self.ber_target)
        
        # see FC-MSQS-2 equation B.47 in Annex B.4 for the following...
        self.k_rin = math.sqrt(2.0/math.pi)*erfinv(0.8)
        # cell Y44
        self.speedup = 1.0 / (1.0 - 1.0E-6*self.br_nominal*self.dcd_dj)
        self.dj_ui = 1.0E-6*self.speedup*self.br_nominal*(self.dj-self.dcd_dj) # cell G9
        # define the length vector, column A
        self.l_stop = 2.0*self.l_target - self.l_start
        self.lnum = 1+int(round((self.l_stop-self.l_start)/self.l_inc))
        self.length = np.linspace(self.l_start,self.l_stop,self.lnum)
        # convenient unit vector used in several calculations
        self.l_1 = np.ones(self.lnum)

#======================================================================+
    def fiber_channel_calc(self):
        """Calculates fiber attenuation, modal bandwidth, and
        chromatic dispersion bandwidth"""

        # calculate fiber attenuation at minimum laser wavelength
        self.alpha = (1.05+(1.0/(0.00094*self.lambda_min))**4)

        # calculate channel insertion loss column C
        self.chil = self.alpha*self.length + self.fiber_conn_loss*self.l_1

        # calculate chromatic dispersion bandwidth
        self.d1 = (0.25*self.fiber_s0*self.lambda_min*(1.0- (self.fiber_u0/self.lambda_min)**4)) # cell P9
        self.d2 = 0.7*self.fiber_s0*self.delta_lambda # cell AB4
        dtot = math.sqrt(self.d1**2 + self.d2**2)
        self.bw_cd = (0.187E6/(self.delta_lambda*dtot) /self.length) # column F

        # calculate modal dispersion bandwidth for each link length
        self.bw_md = self.fib_lbwp / self.length # column G

        # end of GbE10.fiber_channel_calc

    #======================================================================+

    def risetime_calc(self):
        """Calculates the 10%-90% risetimes associated with the transmitter,
        the fiber channel (chromatic disperion, modal dispersion), the
        link receiver, and the reference receiver for transmitter eye
        measurements."""

        # given the transmitter's 20%-80% risetime, and assuming a
        # Gaussian impulse response, calculate the 10%-90% risetime
        # cell G3
        self.tx_1090_rise = 1.518*self.tx_2080_rise  
        
        # calculate the effective risetimes for the fiber channel, given
        # the bandwidths calculated in the previous section, assuming
        # a Gaussian impulse response model
        self.cd_1090_rise = 0.48E6 / self.bw_cd
        self.md_1090_rise = 0.48E6 / self.bw_md

        # calculate the risetime for the link receiver, given its
        # bandwidth and assuming a single pole impulse response
        # Cell T7
        self.rx_1090_rise = 0.329E6/self.rx_bw

        # calculate the risetime for the test receiver used for transmitter
        # eye displays, given its bandwidth and assuming a single pole
        # response
        self.rx_txeye_1090_rise = 0.329E6 / self.txeye_rx_bw

        # calculate Te from column H and Tc from column I
        tr_tx_2 = self.tx_1090_rise**2*self.l_1
        tr_rx_2 = self.rx_1090_rise**2*self.l_1
        tr_cd_2 = np.square(self.cd_1090_rise)
        tr_md_2 = np.square(self.md_1090_rise)
        self.te = np.sqrt(tr_cd_2 + tr_md_2 + tr_tx_2) # column H
                                                  
        self.tc = np.sqrt(tr_cd_2 + tr_md_2 + tr_tx_2 + tr_rx_2) # column I
        

        # end of GbE10..risetime_calc
    #======================================================================+
    def isi_calc(self):

        """ Calculates the eye opening at the center and at the corners,
        needed for various penalty calculations"""

        arg1 = erfinv(0.8)
        arg2 = arg1/(1.0E-6*self.speedup*self.br_nominal)
        #print('arg2 : ', arg2)

        # calculate center eye opening with no additional impairments
        self.isi_center = erf(arg2/self.tc) # column Z

        # calculate center eye opening with residual DJ (DJ - DCD)
        # impairment                              # column AD
        arg3 = erf(arg2*(1.0+self.dj_ui)/self.tc) # column AI
        arg4 = erf(arg2*(1.0-self.dj_ui)/self.tc) # column AJ

        #print(arg3, arg4)
        
        self.isi_dj_center = arg3 + arg4 - self.l_1 # column AD

        # calculate eye closing induced by interferometric effects from
        # link end reflections
        mean_reflection = math.pow(10.0,0.05*(self.rx_reflection + self.tx_reflection)) # cell AB5
        er_lin = math.pow(10.0,0.1*self.er_dB_min) # cell AB7
        arg5 = np.sqrt(2.0*er_lin*self.isi_dj_center*(er_lin-1.0) + (er_lin+1.0)*self.l_1)
        arg6 = np.divide(arg5,self.isi_dj_center)
        arg7 = (2.0*self.ref_nf*np.power(10.0,-0.1*self.chil)*mean_reflection)
        self.isi_reflection = self.l_1-np.multiply(arg6,arg7)

        # calculate center eye opening with both residual DJ and reflection
        # degradations included
        self.isi_dj_closed = np.multiply(self.isi_dj_center,
        self.isi_reflection)                        # column AA

        # calculate eye opening at the corners with no additional impairments
        eff_rx_eye = 2.0*(0.5-self.X2)*self.speedup
        arg8 = erf(arg2*(1.0+eff_rx_eye)/self.tc) # column AE
        arg9 = erf(arg2*(1.0-eff_rx_eye)/self.tc) # column AF
        self.isi_corners = arg8 + arg9 - self.l_1 # column AB

        # calculate eye opening at the corners with residual DJ impairment
        arg10 = erf(arg2*(1.0+eff_rx_eye+self.dj_ui)/self.tc) # column AG
        arg11 = erf(arg2*(1.0-eff_rx_eye-self.dj_ui)/self.tc) # column AH
        self.isi_dj_corners = arg10 + arg11 - self.l_1        # column AC
        
        # end of GbE10.isi_calcdef isi_calc(self):
    #======================================================================+

    def penalty_calc(self):
        """Calculates the various link budget penalties"""
        self.p_budget = (self.tx_oma_min
                         - self.rx_unstressed_sensitivity
                         - self.fiber_conn_loss)*self.l_1

        # fiber attenuation,
        self.p_atten = self.alpha*self.length # column B

        # calculate bandwidth for RIN test (exclude transmitter)
        rin_inverse_bw = np.sqrt(np.square(1.0/self.bw_cd)
                                 + np.square(1.0/self.bw_md)
                                 + (0.477/(self.rx_bw**2))*self.l_1)

        rin_bw = 1.0 / rin_inverse_bw

        # v_rin,
        self.v_rin = (0.7E6*(self.rin_test_isi**2)
                      *math.pow(10.0,0.1*self.rin)*rin_bw) # column AK

        # Prin,
        self.p_rin = -10.0*np.log10(np.sqrt(1.0-np.multiply(self.v_rin,
                                                            np.square(self.Q/self.isi_dj_closed)))) # column R
        self.beta = (3.14159E-6*self.speedup*self.br_nominal *self.delta_lambda*self.d1*self.length) # column O
        self.sigma_mpn = (self.k_mpn/math.sqrt(2.0)*(self.l_1 -np.exp(-np.square(self.beta)))) # column P
        self.p_mpn = (-10.0*np.log10(np.sqrt(1.0- (self.Q**2)*np.square(self.sigma_mpn)))) # column Q
        self.p_blw = (-10.0*math.log10(math.sqrt(1.0- (self.Q*self.sigma_blw)**2))*self.l_1) # cell T13
        self.p_reflection = -10.0*np.log10(self.isi_reflection) # column N


        self.v_mn = (((1.0-math.pow(10.0,-0.2*self.pmn))/ (self.Q)**2)*self.l_1) # cell AG7
        self.p_isi_center = (-10.0*np.log10(2.0*self.isi_center - self.l_1)) # column J
        self.p_isi_corners = (-10.0*np.log10(self.isi_corners) - self.p_isi_center) # column K
        self.p_dj_center = (-10.0*np.log10(self.isi_dj_closed) - self.p_isi_center) # column L
        self.p_dj_corners = (-10.0*np.log10(self.isi_dj_corners) -self.p_isi_center -self.p_isi_corners) # column M

        # calculate the "cross" penalty contribution, column S
        arg1 = (self.sigma_blw**2 + self.v_rin)/np.square(self.isi_dj_closed)
        arg2 = self.l_1 - (self.Q**2)*(arg1 + self.v_mn
        + np.square(self.sigma_mpn))
        arg3 = -10.0*np.log10(np.multiply(self.isi_dj_closed,np.sqrt(arg2)))



        self.p_cross_center = (                         # column S
                                  arg3
                                - self.p_blw            # cell T13
                                - self.p_isi_center     # column J
                                - self.p_dj_center      # column L
                                - self.p_mpn            # column Q
                                - self.p_reflection     # column N
                                - self.p_rin            # column R
                                - self.pmn*self.l_1)    # cell G13
        #print('p_isi_center: ', self.p_isi_center)

        # calculate the total power budget evaluated at the center of the eye
        self.p_total_center = (                         # column T
                                self.p_isi_center       # column J
                                + self.p_dj_center      # column L
                                + self.p_atten          # column B
                                + self.p_mpn            # column Q
                                + self.p_reflection     # column N
                                + self.p_rin            # column R
                                + self.p_cross_center   # column S
                                + self.pmn*self.l_1)    # cell G13
        # calculate the total power budget evaluated at the corner of the eye
        self.p_total_corners = (
                                self.p_isi_center       # column J
                                + self.p_isi_corners    # column K
                                + self.p_atten          # column B
                                + self.p_mpn            # column Q
                                + self.p_reflection     # column N
                                + self.p_rin            # column R
                                + self.p_cross_center   # column S
                                + self.pmn*self.l_1     # cell G13
                                + self.p_dj_corners)    # column M

        self.margin = self.p_budget - self.p_total_center

        print("Margin : ", self.margin)

        print("Length : ", self.length)



        # end of GbE10.penalty_calc
        #======================================================================+


    def eye_calc(self):

        """Calculates the eye diagrams for the link at target reach
        and for the transmitter at test"""
        # define the eye time vector scaled to UI, dimensionless
        tnum = (1+int(round((self.eye_time_high-self.eye_time_low) / self.eye_time_step)))
        self.time = np.linspace(self.eye_time_low, # column Z
                                self.eye_time_high,
                                tnum)
        # convenience vector
        t_1 = np.ones(tnum)
        # column AA
        self.time_eff = 0.5*t_1 + self.speedup*(self.time - 0.5*t_1)
        arg1 = 2.0*erfinv(0.8)
        arg2 = arg1/(1.0E-6*self.speedup*self.br_nominal)
        
        # calculate eye profiles for link at target reach
        T_c_link = self.tc[self.lnum//2]
        arg3 = arg2*self.time_eff/T_c_link # column AP
        arg4 = arg2*(1.0 - self.time_eff)/T_c_link # column AQ
        self.link_011 = 0.5*t_1 + 0.5*erf(arg3) # column AR
        self.link_110 = 0.5*t_1 + 0.5*erf(arg4) # column AS
        self.link_010 = self.link_011 + self.link_110 - t_1 # column AT
        self.link_101 = t_1 - self.link_010 # column AW
        self.link_001 = t_1 - self.link_110 # column AU
        self.link_100 = t_1 - self.link_011 # column AV

        # calculate eye profiles for transmitter eye test configuration
        T_c_test = math.sqrt(self.tx_1090_rise**2
        + (0.329E6/self.txeye_rx_bw)**2)
        arg5 = arg2*self.time_eff/T_c_test
        arg6 = arg2*(t_1 - self.time_eff)/T_c_test
        self.test_011 = 0.5*t_1 + 0.5*erf(arg5) # column AR
        self.test_110 = 0.5*t_1 + 0.5*erf(arg6) # column AS
        self.test_010 = self.test_011 + self.test_110 - t_1 # column AT
        self.test_101 = t_1 - self.test_010
        self.test_001 = t_1 - self.test_110
        self.test_100 = t_1 - self.test_011


        
# end of GbE10.eye_calc
#======================================================================+
# |
# End of file Gb10E_support.py |
# |
#======================================================================+



if __name__ == '__main__':
    
    import matplotlib.pyplot as plt

    g = GbE10()

    fig, axs = plt.subplots(2)
    fig.suptitle('10Gbe Plots')

    
    axs[0].plot (g.length, g.p_budget,label='budget')
    axs[0].plot (g.length, g.p_total_center,label='P_total (central)')
    axs[0].plot (g.length, g.p_total_corners,label='P_total(corners)')
    axs[0].plot (g.length, g.p_isi_center,label='P_isi(central)')
    axs[0].plot (g.length, g.p_atten,label='P_atten')
    axs[0].plot (g.length, g.p_cross_center,label='P_cross(central)' )
    axs[0].legend(loc='upper left')


    axs[1].plot (g.time, g.link_001 , label="erf2' ")
    axs[1].plot (g.time, g.link_010, label="oio ")
    axs[1].plot (g.time, g.link_011,  label="erf1 ")
    axs[1].plot (g.time, g.link_100, label="erf1' ")
    axs[1].plot (g.time, g.link_101, label="ioi ")
    axs[1].plot (g.time, g.link_110, label="erf2 ")
    axs[1].plot (g.time, g.test_010, label="oio ")
    axs[1].plot (g.time, g.test_101 , label="ioi ")
    axs[1].legend(loc='upper left')

    plt.show()







